import fsrl
# print(fsrl.__version__)
from fsrl.agent.cpo_agent import CPOAgent
from fsrl.agent.cvpo_agent import CVPOAgent
import copy
import gymnasium as gym
from gymnasium import spaces
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from fsrl.utils import TensorboardLogger
from tianshou.env import DummyVectorEnv
from fsrl.data import FastCollector
import matplotlib.pyplot as plt


def load_data(filepath):
    wb = load_workbook(filepath)
    sheet = wb.active
    data = []
    date = []
    for row in sheet.iter_rows(min_row=2):  # assuming the first row is headers
        for cell in row:
            if isinstance(cell.value, datetime):
                # Convert datetime to a string or another preferred format
                date.append(cell.value)
            elif cell.value is not None:
                data.append(np.float32(cell.value) / 10000)
    min_value = 0
    max_value = 0.2
    data = np.array(data)
    normalized_array = np.interp(data, (data.min(), data.max()), (min_value, max_value))
    return data, date


class EVChargingEnv(gym.Env):
    def __init__(self):
        super(EVChargingEnv, self).__init__()
        self.spec = gym.envs.registration.EnvSpec("EVCharging-v0")
        self.spec.max_episode_steps = 1024
        self.spec.reward_threshold = 0.0
        self.spec.name = "EVCharging"
        self.departure_time = 0
        self.arrival_time = 0
        self.time_step = 0
        self.time_set = False
        self.max_episode_steps = 1024
        self.max_battery_capacity = 24.0  # kWh
        self.min_battery_capacity = 2.4  # kWh
        self.target_battery_capacity = 21.0  # kWh
        self.max_charge_rate = 6.0  # kWh
        self.max_discharge_rate = -6.0  # kWh
        self.begin = 24
        self.data, self.date = load_data("intradayprice.xlsx")
        self.single_date = 0
        self.seed = 0
        # 状态空间：EV电池能量和过去24小时电价
        self.observation_space = spaces.Box(low=np.array([self.min_battery_capacity] + [0.0] * 24, dtype=np.float64),
                                            high=np.array([self.max_battery_capacity] + [np.inf] * 24,
                                                          dtype=float), shape=(25,),
                                            dtype=np.float64)

        # 动作空间：充电或放电能量
        self.action_space = spaces.Box(low=self.max_discharge_rate, high=self.max_charge_rate, shape=(1,),
                                       dtype=np.float64)

        # 初始化状态
        self.state = None
        self.reset()

    def reset(self):
        if self.begin == self.data.size:
            self.begin = 24
        info = {}
        # 初始化EV电池能量和过去24小时电价（假设为随机值）
        if (self.begin % 24) != 0:
            self.begin -= self.begin % 24

        self.arrival_time = int(round(np.random.normal(18, 1)))
        self.arrival_time = min(self.arrival_time, 21)
        self.arrival_time = max(self.arrival_time, 15)
        if self.begin + self.arrival_time >= self.data.size or self.data.size - (self.begin + self.arrival_time) < (
                24 - self.arrival_time + self.departure_time):
            self.begin = 24
        self.departure_time = int(round(np.random.normal(8, 1)))
        self.departure_time = min(self.departure_time, 11)
        self.departure_time = max(self.departure_time, 6)
        initial_battery = np.random.normal(12, 2.4)
        initial_battery = min(initial_battery, 19.2)
        initial_battery = max(initial_battery, 4.8)
        self.begin += self.arrival_time
        # print(self.begin)
        # print(self.begin)
        past_24h_prices = self.data[self.begin - 24:self.begin]  # 随机电价，实际应用中可用实际数据
        self.state = np.array([initial_battery] + list(past_24h_prices))
        return self.state, info

    def step_test(self, action, charging_price: float, tr):
        # 检查是否完成（假设一天结束）
        dw = False  # 在此示例中，不定义特定完成条件
        # tr = False
        battery_energy: float = self.state[0]
        past_24h_prices = self.state[1:]

        # 当前时间步的电价（假设为随机值，实际应用中应为实际数据）
        current_charging_price = charging_price
        current_discharging_price = charging_price

        # 更新电池能量
        new_battery_energy: float = battery_energy + action[0]
        # 返回约束（计算差异）
        # cost function
        constraint_violation = 0.0
        if tr:
            constraint_violation = abs(new_battery_energy - self.target_battery_capacity)
        elif new_battery_energy > self.max_battery_capacity:
            constraint_violation = new_battery_energy - self.max_battery_capacity
        elif new_battery_energy < self.min_battery_capacity:
            constraint_violation = self.min_battery_capacity - new_battery_energy
        # 计算奖励
        if action[0] < 0:
            reward = -action[0] * current_charging_price - 0.1 * constraint_violation
        else:
            if new_battery_energy < self.min_battery_capacity:
                reward = 0.0 - 0.1 * constraint_violation
            else:
                reward = -action[0] * current_discharging_price - 0.1 * constraint_violation
        # 计算能量损失
        if action[0] > 0:
            energy_loss = action[0] * 0.02  # 假设充电效率为98%
        else:
            energy_loss = action[0] * -0.02  # 假设放电效率为98%

        new_battery_energy -= energy_loss
        if new_battery_energy > self.max_battery_capacity:
            new_battery_energy = self.max_battery_capacity
        elif new_battery_energy < self.min_battery_capacity:
            new_battery_energy = self.min_battery_capacity

        # 更新状态
        new_state = np.array([new_battery_energy] + list(past_24h_prices[1:]) + [current_charging_price])
        self.state = new_state

        info = {'constraint_violation': constraint_violation, 'reward': reward, 'price': current_charging_price}
        return new_state, reward, dw, tr, info

    def step(self, action):
        # 检查是否完成（假设一天结束）
        dw = False  # 在此示例中，不定义特定完成条件
        tr = False
        #print(action[0])
        if self.arrival_time > 12:

            if self.single_date > (24 - self.arrival_time + self.departure_time):
                self.single_date = 0
                # get_next_arrival_time(self.date[self.begin])
                tr = True
        else:
            if self.single_date > (self.departure_time - self.arrival_time):
                self.single_date = 0
                # get_next_arrival_time(self.date[self.begin])
                tr = True
        battery_energy: float = self.state[0]
        past_24h_prices = self.state[1:]
        # print(past_24h_prices)
        # print(self.begin)
        # 当前时间步的电价（假设为随机值，实际应用中应为实际数据）
        current_charging_price = self.data[self.begin]
        current_discharging_price = self.data[self.begin]
        new_battery_energy = 0.0
        # 更新电池能量
        new_battery_energy: float = battery_energy + action[0]
        # 返回约束（计算差异）
        # cost function
        constraint_violation = 0.0
        if tr:
            constraint_violation = abs(new_battery_energy - self.target_battery_capacity)
        elif new_battery_energy > self.max_battery_capacity:
            constraint_violation = new_battery_energy - self.max_battery_capacity
        elif new_battery_energy < self.min_battery_capacity:
            constraint_violation = self.min_battery_capacity - new_battery_energy
        # 计算奖励
        reward = 0.0
        if action[0] < 0:
            reward = -action[0] * current_charging_price - 0.05 * constraint_violation
        else:
            if new_battery_energy < self.min_battery_capacity:
                reward = 0.0 - 0.05 * constraint_violation
            else:
                reward = -action[0] * current_discharging_price - 0.05 * constraint_violation
        # 计算能量损失
        if action[0] > 0:
            energy_loss = action[0] * 0.02  # 假设充电效率为98%
        else:
            energy_loss = action[0] * -0.02  # 假设放电效率为98%
        reward = np.nan_to_num(reward)
        constraint_violation = np.nan_to_num(constraint_violation)
        new_battery_energy -= energy_loss
        if new_battery_energy > self.max_battery_capacity:
            new_battery_energy = self.max_battery_capacity
        elif new_battery_energy < self.min_battery_capacity:
            new_battery_energy = self.min_battery_capacity
        # 更新状态
        new_battery_energy = np.nan_to_num(new_battery_energy)
        new_state = np.array([new_battery_energy] + list(past_24h_prices[1:]) + [current_charging_price],
                             dtype=np.float64)
        self.state = new_state

        # print(current_charging_price,",",reward,",",constraint_violation)
        self.begin += 1
        self.single_date += 1
        if self.begin == self.data.size:
            tr = True
        # if tr and constraint_violation < 0.1:
        #     dw = True
        if tr and constraint_violation < 0.5:
            constraint_violation = 0.0
        info = {'cost': constraint_violation}
        # print(new_state, reward)
        return new_state, reward, dw, tr, info

    def render(self, mode='human'):
        if not hasattr(self, 'fig'):
            self.fig, self.ax1 = plt.subplots()
            self.ax2 = self.ax1.twinx()
            self.ax1.set_xlabel('Time Step')
            self.ax1.set_ylabel('Battery Energy (kWh)', color='b')
            self.ax2.set_ylabel('Price ($/kWh)', color='r')
            self.ax1.set_ylim(self.min_battery_capacity, self.max_battery_capacity)
            self.ax2.set_ylim(0, max(self.data) + 1)  # 假设电价不会超过最大值+1
            self.battery_line, = self.ax1.plot([], [], 'b-', label='Battery Energy')
            self.price_line, = self.ax2.plot([], [], 'r-', label='Current Price')
            self.ax1.legend(loc='upper left')
            self.ax2.legend(loc='upper right')
            plt.title('EV Charging Environment')
            plt.show(block=False)

        # 获取当前状态
        battery_energy = self.state[0]
        current_price = self.state[-1]

        # 更新电池能量和电价的折线图
        self.battery_line.set_xdata(list(range(self.single_date + 1)))
        self.battery_line.set_ydata([self.state[0]] * (self.single_date + 1))

        self.price_line.set_xdata(list(range(self.single_date + 1)))
        self.price_line.set_ydata([self.state[-1]] * (self.single_date + 1))

        # 重新绘制图表
        self.ax1.relim()
        self.ax1.autoscale_view()
        self.ax2.relim()
        self.ax2.autoscale_view()
        self.fig.canvas.draw()
        self.fig.canvas.flush_events()

        plt.pause(0.01)


def main():
    # EnvName = ['Pendulum-v1', 'LunarLanderContinuous-v2', 'Humanoid-v4', 'HalfCheetah-v4', 'BipedalWalker-v3',
    #            'BipedalWalkerHardcore-v3']
    # BrifEnvName = ['PV1', 'LLdV2', 'Humanv4', 'HCv4', 'BWv3', 'BWHv3']

    env = EVChargingEnv()
    eval_env = copy.deepcopy(env)
    logger = TensorboardLogger("logs", log_txt=True, name="EVChargingEnv")
    # cpo_agent = CPOAgent(env=env, logger=logger, cost_limit=0.1, lr=2e-3, l2_reg=0.1, reward_normalization=True)
    cvpo_agent = CVPOAgent(env=env, logger=logger, hidden_sizes=(100, 200), cost_limit=1e-5)
    training_num, testing_num = 5, 1
    # collector = FastCollector(agent.policy, env)
    train_envs = DummyVectorEnv([lambda: EVChargingEnv() for _ in range(training_num)])
    test_envs = DummyVectorEnv([lambda: EVChargingEnv() for _ in range(testing_num)])

    # agent.policy.train()
    step_per_epoch = int(1e5)
    cvpo_agent.learn(train_envs=train_envs, test_envs=test_envs, reward_threshold=-0.1, step_per_epoch=step_per_epoch,
                     epoch=3000, testing_num=365)
    # cpo_agent.learn(train_envs=train_envs, test_envs=test_envs, reward_threshold=0.0, batch_size=4096, testing_num = 365)
    # result = collector.collect(n_episode=10, render=True)


if __name__ == '__main__':
    main()
